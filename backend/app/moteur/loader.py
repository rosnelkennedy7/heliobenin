import openpyxl
import os
import glob


def _find_excel(keyword: str):
    data_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'data')
    for pattern in [os.path.join(data_dir, '*.xlsx'), os.path.join(data_dir, '**/*.xlsx')]:
        for f in glob.glob(pattern):
            name = os.path.basename(f).lower()
            if keyword in name:
                return f
    return None


EXCEL_PATH           = _find_excel('equipement') or _find_excel('equipment')
EXCEL_APPAREILS_PATH = _find_excel('appareils')
EXCEL_PRIX_PATH      = _find_excel('prix')

if not EXCEL_PATH:
    raise FileNotFoundError("HélioBénin Equipement.xlsx introuvable dans data/")


def _parse_plage(s):
    try:
        clean = str(s).replace('V', '').strip()
        parts = clean.split('-')
        return float(parts[0]), float(parts[-1])
    except Exception:
        return 0.0, 500.0


def _load_sheet_from(path: str, sheet_name: str) -> list:
    if not path:
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(v is not None for v in row)
    ]


def load_sheet(sheet_name: str) -> list:
    return _load_sheet_from(EXCEL_PATH, sheet_name)


def _normalize_panneaux(rows):
    result = []
    for r in rows:
        try:
            result.append({
                'marque': r.get('Marque'),
                'modele': r.get('Modèle'),
                'type': r.get('Type'),
                'puissance': float(r.get('P (Wc)') or 0),
                'tension_nominale': int(r.get('Tension nominale (V)') or 0),
                'voc': float(r.get('Voc (V)') or 0),
                'isc': float(r.get('Isc (A)') or 0),
                'vmp': float(r.get('Vmp (V)') or 0),
                'imp': float(r.get('Imp (A)') or 0),
                'rendement': float(r.get('Rendement (%)') or 0),
            })
        except Exception:
            continue
    return result


def _normalize_onduleurs_aio(rows):
    result = []
    for r in rows:
        try:
            mppt_min, mppt_max = _parse_plage(r.get('Plage MPPT (V)', '0-500'))
            result.append({
                'marque': r.get('Marque'),
                'modele': r.get('Modèle'),
                'puissance': int(r.get('Puissance (W)') or 0),
                'tension_sortie': int(r.get('Tension sortie AC (V)') or 230),
                'usys': int(r.get('Tension système (V)') or 48),
                'courant_charge_bat': float(r.get('Courant charge bat. (A)') or 0),
                'mppt_min': mppt_min,
                'mppt_max': mppt_max,
                'courant_mppt_max': float(r.get('Courant MPPT max (A)') or 999),
                'pv_max': float(r.get('Puissance PV max (W)') or 0),
                'nb_mppt': int(r.get('Nb MPPT') or 1),
                'rendement': float(r.get('Rendement (%)') or 97),
            })
        except Exception:
            continue
    return result


def _normalize_batteries(rows):
    result = []
    for r in rows:
        try:
            result.append({
                'marque': r.get('Marque'),
                'modele': r.get('Modèle'),
                'technologie': r.get('Technologie'),
                'tension': int(r.get('Tension (V)') or 0),
                'capacite': float(r.get('Capacité (Ah)') or 0),
                'energie': float(r.get('Énergie (kWh)') or 0),
                'dod': float(r.get('DoD (%)') or 90),
                'rendement': float(r.get('Rendement (%)') or 95),
                'courant_decharge_max': float(r.get('Courant décharge max (A)') or 0),
            })
        except Exception:
            continue
    return result


def _normalize_regulateurs(rows):
    result = []
    for r in rows:
        try:
            result.append({
                'marque': r.get('Marque'),
                'modele': r.get('Modèle'),
                'type': r.get('Type'),
                'courant_max': float(r.get('Courant max (A)') or 0),
                'tension_systeme': str(r.get('Tension système (V)') or ''),
                'plage_pv': str(r.get('Plage tension PV (V)') or ''),
            })
        except Exception:
            continue
    return result


def _normalize_appareils(rows):
    result = []
    for i, r in enumerate(rows, 1):
        try:
            nom = str(r.get('Désignation') or '').strip()
            if not nom:
                continue
            id_ = i
            for key, val in r.items():
                kstr = str(key)
                if kstr.startswith('N') and len(kstr) <= 3 and val is not None:
                    try:
                        id_ = int(val)
                    except (ValueError, TypeError):
                        pass
                    break
            result.append({
                'id': id_,
                'nom': nom,
                'categorie': str(r.get('Catégorie') or '').strip(),
                'puissance': float(r.get('Puissance (W)') or 0),
                'typeCharge': str(r.get('Type de charge') or '').strip(),
                'facteurPointe': float(r.get('Facteur pointe') or 1.0),
            })
        except Exception:
            continue
    return result


def load_all_equipements() -> dict:
    return {
        'panneaux':             _normalize_panneaux(load_sheet('Panneaux Solaires')),
        'batteries':            _normalize_batteries(load_sheet('Batteries')),
        'onduleurs_aio':        _normalize_onduleurs_aio(load_sheet('Onduleurs All-in-One')),
        'onduleurs_hybrides':   load_sheet('Onduleurs Hybrides'),
        'onduleurs_classiques': load_sheet('Onduleurs Classiques'),
        'regulateurs':          _normalize_regulateurs(load_sheet('Régulateurs')),
        'disjoncteurs':         load_sheet('Disjoncteurs'),
        'differentiels':        load_sheet('Différentiels AC'),
        'fusibles':             load_sheet('Fusibles DC'),
        'porte_fusibles':       load_sheet('Porte-fusibles'),
        'parafoudres':          load_sheet('Parafoudres'),
    }


def load_appareils() -> list:
    rows = _load_sheet_from(EXCEL_APPAREILS_PATH, 'Appareils')
    return _normalize_appareils(rows)


def load_prix() -> dict:
    if not EXCEL_PRIX_PATH:
        return {}
    sheets = ['Protection & Câble', 'Onduleurs All-in-One', 'Batteries LiFePO4', 'Panneaux Solaires']
    return {
        sheet: [{str(k): v for k, v in row.items()} for row in _load_sheet_from(EXCEL_PRIX_PATH, sheet)]
        for sheet in sheets
    }
